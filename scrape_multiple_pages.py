from bs4 import BeautifulSoup
import requests
import openpyxl         #for saving data in excel

# Create an Excel workbook and set up the sheet
excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active      # Use the active sheet
sheet.title = 'Top Rated Books on Amazon'
print(excel.sheetnames)
sheet.append(['Book Name', 'Author Name', 'Rating', 'Price'])  # Naming our columns

# Function to fetch books from a specific page
def fetch_books_from_page(page_number):
    url = f'https://www.amazon.in/gp/bestsellers/books?page={page_number}'
    source = requests.get(url)
    source.raise_for_status()
    soup = BeautifulSoup(source.text, 'html.parser')
    return soup

# Function to extract book data from the soup
def extract_books(soup):
    book_list = []
    books = soup.find_all('div', class_='p13n-sc-uncoverable-faceout')
    for book in books:
        name=book.find('div',class_="_cDEzb_p13n-sc-css-line-clamp-1_1Fn1y").get_text(strip=True)
        author=book.find('div',class_="a-row a-size-small").get_text(strip=True)
        rating=book.find('div',class_="a-icon-row").get_text(strip=True).split(' ')[0]
        price = book.find('span', class_='p13n-sc-price').get_text(strip=True)
        
        book_list.append((name, author, rating, price))
    return book_list

# Main scraping loop
all_books = []
try:
    for page_number in range(1, 9):  # Adjust the range based on the number of pages you want to scrape
        soup = fetch_books_from_page(page_number)
        books = extract_books(soup)
        all_books.extend(books)
        print(f"Fetched {len(books)} books from page {page_number}")

    # Append book data to the Excel sheet
    for book in all_books:
        sheet.append(book)

except Exception as e:
    print(e)

# Save the Excel file
excel.save('BestSeller Books of Amazon.csv')
